from __future__ import annotations

import csv
import hashlib
import json
import sqlite3
import logging
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

# Setup logging configuration
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

# -----------------------------
# Configuration
# -----------------------------
# Standard input configuration for all known vendors
EXPECTED_INPUTS = [
    {
        "source_vendor": "dental",
        "file_name": "dental_provider.xlsx",
        "format": "xlsx",
        "sheet_name": "eligibility",
    },
    {
        "source_vendor": "vision",
        "file_name": "vision_provider.csv",
        "format": "csv",
        "delimiter": ",",
        "has_header": True,
    },
    {
        "source_vendor": "medical_a",
        "file_name": "medical_provider_a.csv",
        "format": "csv",
        "delimiter": ",",
        "has_header": True,
    },
    {
        "source_vendor": "medical_b",
        "file_name": "medical_provider_b.txt",
        "format": "txt",
        "delimiter": "|",
        "has_header": True,
    },
    {
        "source_vendor": "medical_c",
        "file_name": "medical_provider_c.jsonl",
        "format": "jsonl",
    },
]

# -----------------------------
# Data models
# -----------------------------
@dataclass(frozen=True)
class ManifestFileEntry:
    source_vendor: str
    source_file: str
    relative_path: str
    size_bytes: int
    modified_time_utc: str
    sha256: str
    row_count_read: int
    status: str  # "success" | "failed"
    error: str | None

@dataclass(frozen=True)
class StagingManifest:
    load_run_id: str
    ingested_at_utc: str
    input_dir: str
    files: list[ManifestFileEntry]

# -----------------------------
# Helpers: File Metadata & IDs
# -----------------------------
def _utc_now_iso() -> str:
    """Return current UTC timestamp in ISO format (no microseconds)."""
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def generate_load_run_id() -> str:
    """Generate a unique load_run_id for each pipeline run."""
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"{ts}_{uuid4().hex[:8]}"

def compute_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    """Compute SHA256 checksum for a file in a streaming manner."""
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()

def file_modified_time_utc(path: Path) -> str:
    """Return file modified time as UTC ISO timestamp."""
    mtime = path.stat().st_mtime
    return datetime.fromtimestamp(mtime, tz=timezone.utc).replace(microsecond=0).isoformat()

# -----------------------------
# Helpers: Row Counting Logic
# -----------------------------
def count_rows_csv_like(path: Path, delimiter: str, has_header: bool = True) -> int:
    """Count rows in CSV/TXT files. Header is excluded if has_header=True."""
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        count = 0
        for i, row in enumerate(reader, start=1):
            if not row:  # Skip empty lines
                continue
            if has_header and i == 1:
                continue
            count += 1
    return count

def count_rows_jsonl(path: Path) -> int:
    """Count non-empty JSON Lines."""
    count = 0
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                count += 1
    return count

def count_rows_xlsx(path: Path, sheet_name: str | None = None) -> int:
    """Count data rows in Excel, assuming row 1 is header."""
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            count += 1
    wb.close()
    return count

def count_rows_by_format(path: Path, fmt: str, meta: dict[str, Any]) -> int:
    """Dispatch row counting based on file extension/format."""
    if fmt in ("csv", "txt"):
        return count_rows_csv_like(
            path=path,
            delimiter=meta.get("delimiter", ","),
            has_header=meta.get("has_header", True),
        )
    if fmt == "jsonl":
        return count_rows_jsonl(path)
    if fmt == "xlsx":
        return count_rows_xlsx(path, sheet_name=meta.get("sheet_name"))
    raise ValueError(f"Unsupported format: {fmt}")

# ----------------------------------------------------------
# Database Initialization: The "Clean Slate" Strategy
# ----------------------------------------------------------
def init_db(root: Path):
    """
    Initializes the SQLite database.
    Drops existing tables to ensure a clean state for the new run.
    """
    output_dir = root / "output"
    sql_dir = root / "sql"
    db_path = output_dir / "warehouse.db"

    output_dir.mkdir(parents=True, exist_ok=True)

    # Establish connection
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Drop existing tables to prevent data pollution from previous runs
    tables_to_drop = ["silver_members", "raw_staging_payload", "raw_staging"]
    for table in tables_to_drop:
        cur.execute(f"DROP TABLE IF EXISTS {table}")
        logger.info(f"Dropped table (if existed): {table}")

    # Recreate tables using DDL scripts from the /sql folder
    ddl_files = [
        "01_create_raw_staging.sql",
        "02_create_raw_staging_payload.sql",
        "03_create_silver_members.sql"
    ]

    for ddl_file in ddl_files:
        ddl_path = sql_dir / ddl_file
        if ddl_path.exists():
            with open(ddl_path, 'r') as f:
                cur.executescript(f.read())
            logger.info(f"Executed DDL: {ddl_file}")
        else:
            logger.warning(f"DDL script missing: {ddl_path}")

    conn.commit()
    conn.close()
    logger.info(f"Database initialized and clean: {db_path}")
    return db_path

# -----------------------------
# Stage 0: Manifest Generation
# -----------------------------
def build_staging_manifest(root: Path, load_run_id: str, require_non_empty: bool = True) -> Path:
    """
    Scans input files, computes checksums and row counts, and writes a JSON manifest.
    This manifest serves as a technical log for the current ingestion run.
    """
    input_dir = root / "input"
    output_dir = root / "output"
    manifests_dir = output_dir / "manifests"

    output_dir.mkdir(parents=True, exist_ok=True)
    manifests_dir.mkdir(parents=True, exist_ok=True)

    entries: list[ManifestFileEntry] = []

    for spec in EXPECTED_INPUTS:
        vendor = spec["source_vendor"]
        file_name = spec["file_name"]
        fmt = spec["format"]
        file_path = input_dir / file_name

        try:
            if not file_path.exists():
                raise FileNotFoundError(f"Missing expected input file: {file_path}")

            # Collect metadata
            size = file_path.stat().st_size
            mtime = file_modified_time_utc(file_path)
            sha = compute_sha256(file_path)
            rows = count_rows_by_format(file_path, fmt, spec)

            if require_non_empty and rows == 0:
                raise ValueError(f"File contains no data rows: {file_name}")

            entries.append(
                ManifestFileEntry(
                    source_vendor=vendor,
                    source_file=file_name,
                    relative_path=str(file_path.relative_to(root)),
                    size_bytes=size,
                    modified_time_utc=mtime,
                    sha256=sha,
                    row_count_read=rows,
                    status="success",
                    error=None,
                )
            )

        except Exception as e:
            logger.error(f"Error processing {file_name}: {e}")
            entries.append(
                ManifestFileEntry(
                    source_vendor=vendor,
                    source_file=file_name,
                    relative_path=str(file_path.relative_to(root)) if file_path.exists() else f"input/{file_name}",
                    size_bytes=file_path.stat().st_size if file_path.exists() else 0,
                    modified_time_utc=file_modified_time_utc(file_path) if file_path.exists() else "",
                    sha256=compute_sha256(file_path) if file_path.exists() else "",
                    row_count_read=0,
                    status="failed",
                    error=str(e),
                )
            )

    # Wrap as manifest object
    manifest = StagingManifest(
        load_run_id=load_run_id,
        ingested_at_utc=_utc_now_iso(),
        input_dir=str(input_dir.relative_to(root)),
        files=entries,
    )

    # Save to JSON
    manifest_path = manifests_dir / f"manifest_{load_run_id}.json"
    latest_path = output_dir / "staging_manifest_latest.json"

    content = json.dumps(asdict(manifest), ensure_ascii=False, indent=2)
    manifest_path.write_text(content, encoding="utf-8")
    latest_path.write_text(content, encoding="utf-8")

    logger.info(f"Manifest generated successfully: {manifest_path}")
    return manifest_path